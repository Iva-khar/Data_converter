from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment, fonts, PatternFill
from openpyxl.utils.cell import get_column_letter
from rest_framework import viewsets

from business_register.filters import FopFilterSet
from business_register.models.fop_models import Fop
from business_register.serializers.fop_serializers import FopSerializer
from data_ocean.views import CachedViewMixin, RegisterViewMixin
from rest_framework.filters import SearchFilter


class FopViewSet(RegisterViewMixin,
                 CachedViewMixin,
                 viewsets.ReadOnlyModelViewSet):
    queryset = Fop.objects.select_related(
        'status', 'authority'
    ).prefetch_related(
        'kveds', 'exchange_data'
    ).all()
    filter_backends = (DjangoFilterBackend, SearchFilter)
    serializer_class = FopSerializer
    filterset_class = FopFilterSet
    search_fields = ('fullname', 'address', 'status__name')

    def list(self, request, *args, **kwargs):
        try:
            if request.GET['export'] == 'xlsx':
                queryset = self.filter_queryset(self.get_queryset())
                export_dict = {
                    'Full Name': 'fullname',
                    'Status': 'status',
                    'Address': 'address',
                    'Registration Date': 'registration_date',
                    'Termination Date': 'termination_date',
                }
                export_file_path = settings.EXPORT_FOLDER + 'fop_{0}.xlsx'.format(
                    datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                )
                worksheet_title = 'FOP'
                export_xlsx(queryset, export_file_path, export_dict, worksheet_title)
                return HttpResponse(export_file_path, content_type="text/plain")
        except:
            return super().list(request, *args, **kwargs)


def export_xlsx(queryset, export_file_path, export_dict, worksheet_title):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title
    worksheet.sheet_properties.tabColor = '0033CCCC'
    worksheet.row_dimensions[1].height = 20
    worksheet.freeze_panes = 'A2'
    for col_num, (column_title, query_field) in enumerate(export_dict.items(), 1):
        row_num = 1
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        worksheet.column_dimensions[get_column_letter(col_num)].width = 30
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = fonts.Font(b=True)
        cell.fill = PatternFill(bgColor='0033CCCC', fill_type="solid")
        for record in queryset:
            sell_value = getattr(record, query_field)
            row_num += 1
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            try:
                cell.value = sell_value
            except:
                cell.value = repr(sell_value)
    return workbook.save(export_file_path)
